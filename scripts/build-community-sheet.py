"""
Build the community-suggestions workbook (data-sources/community-suggestions.xlsx).

Three input tabs for the community to fill in:
  1. Overall Adjustments — every current player/coach/sub card with its current
     overall; community proposes a new one. Columns MIRROR the CSV that
     scripts/apply-overall-review.mjs reads, so a filled tab can be exported back
     to CSV and applied directly.
  2. Special Card Suggestions — propose new special cards.
  3. New Team Suggestions — propose lineups to add (Liquipedia-sourced).

Regenerate the overall baseline first, then this:
  node scripts/apply-overall-review.mjs --export data-sources/overall-review-current.csv
  python scripts/build-community-sheet.py
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(ROOT, "data-sources", "overall-review-current.csv")
OUT_PATH = os.path.join(ROOT, "data-sources", "community-suggestions.xlsx")

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F2A44")      # dark navy
INPUT_FILL = PatternFill("solid", fgColor="FFF3C4")     # soft yellow = "fill here"
EX_FILL = PatternFill("solid", fgColor="EEF2F7")        # grey = example row
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
EX_FONT = Font(name=FONT, size=10, italic=True, color="6B7280")
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, input_cols=()):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = INPUT_FILL if c in input_cols else HEAD_FILL
        if c in input_cols:
            cell.font = Font(name=FONT, bold=True, color="7A5C00", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w


wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet 0: Instructions (Leia-me)
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Leia-me"
ws.sheet_view.showGridLines = False
lines = [
    ("Rocket Draft — Sugestões da comunidade", True, 16),
    ("", False, 11),
    ("Este arquivo tem 3 abas para a comunidade preencher. Colunas em AMARELO são para preencher.", False, 11),
    ("", False, 11),
    ("1) Overall Adjustments", True, 12),
    ("   Toda carta atual (jogador/coach/sub) com o overall atual. Sugira um novo em 'OVR sugerido'.", False, 10),
    ("   Deixe em branco para não mudar. NÃO altere as outras colunas (são a chave de identificação).", False, 10),
    ("", False, 10),
    ("2) Special Card Suggestions", True, 12),
    ("   Proponha novas cartas especiais. Regra: o overall do special deve ser MAIOR que a carta base do jogador.", False, 10),
    ("   Tipos: worlds_mvp, major_mvp, moment, legend, mythic. Raridades: rare, epic, mythic, legendary.", False, 10),
    ("   Efeito simples: um boost direto em 1-2 atributos (ex: +3 mechanics).", False, 10),
    ("", False, 10),
    ("3) New Team Suggestions", True, 12),
    ("   Proponha lineups para adicionar. Inclua a URL da Liquipedia como fonte.", False, 10),
    ("   Org buff: ~ / + / ++ / +++.  Força histórica: elite / strong / solid / underdog.", False, 10),
    ("", False, 11),
    ("Atributos válidos: offense, defense, mechanics, consistency, experience, clutch.", False, 10),
]
for i, (text, bold, size) in enumerate(lines, start=1):
    cell = ws.cell(row=i, column=1, value=text)
    cell.font = Font(name=FONT, bold=bold, size=size, color="1F2A44" if bold else "111827")
ws.column_dimensions["A"].width = 110

# ---------------------------------------------------------------------------
# Sheet 1: Overall Adjustments (from the exported baseline CSV)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Overall Adjustments")
with open(CSV_PATH, encoding="utf-8") as f:
    rows = list(csv.reader(f))
# Per-player "special card" suggestion column (col 12) — the community can flag a
# player who deserves a special and sketch the idea inline; detailed proposals
# (title/type/rarity/OVR/stats) still go on the "Special Card Suggestions" tab.
header = rows[0] + ["Sugestão de carta especial"]
ws.append(header)
for r in rows[1:]:
    ws.append(r)
# "OVR sugerido" (col 9), "Sugestão de line/time" (col 10) and "Sugestão de carta
# especial" (col 12) are the community input cols.
input_cols = (9, 10, 12)
style_header(ws, len(header), input_cols=input_cols)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(header)):
    for cell in row:
        cell.font = BODY_FONT
        cell.border = BORDER
        if cell.column in input_cols:
            cell.fill = INPUT_FILL
set_widths(ws, [9, 16, 24, 16, 7, 9, 8, 9, 11, 22, 30, 34])

# ---------------------------------------------------------------------------
# Sheet 2: Special Card Suggestions
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Special Card Suggestions")
hdr = ["Jogador", "Carta base (ref id, opcional)", "Título sugerido", "Tipo",
       "Raridade", "OVR sugerido", "Atributo(s) do efeito", "Valor", "Stats (opcional)",
       "Momento / justificativa", "Referência (URL)", "Notas"]
ws.append(hdr)
example = ["zen", "zen-team-vitality-2223", "Worlds Champion", "worlds_mvp", "legendary",
           "99", "clutch, mechanics", "3", "", "MVP do Mundial 2022-23",
           "https://liquipedia.net/rocketleague/...", "exemplo — apague esta linha"]
ws.append(example)
style_header(ws, len(hdr), input_cols=tuple(range(1, len(hdr) + 1)))
for cell in ws[2]:
    cell.font = EX_FONT
    cell.fill = EX_FILL
    cell.border = BORDER
set_widths(ws, [14, 26, 22, 13, 12, 11, 22, 8, 18, 30, 32, 26])

# ---------------------------------------------------------------------------
# Sheet 3: New Team Suggestions
# ---------------------------------------------------------------------------
ws = wb.create_sheet("New Team Suggestions")
hdr = ["Org / Time", "Temporada (ex: RLCS 2024)", "Ano", "Região", "Player 1",
       "Player 2", "Player 3", "Coach (opcional)", "Sub (opcional)", "Org buff (~/+/++/+++)",
       "Força histórica", "Liquipedia URL", "Notas"]
ws.append(hdr)
example = ["Team Falcons", "RLCS 2024", "2024", "MENA", "trk511", "kiileerz", "rw9",
           "", "", "++", "strong", "https://liquipedia.net/rocketleague/Team_Falcons",
           "exemplo — apague esta linha"]
ws.append(example)
style_header(ws, len(hdr), input_cols=tuple(range(1, len(hdr) + 1)))
for cell in ws[2]:
    cell.font = EX_FONT
    cell.fill = EX_FILL
    cell.border = BORDER
set_widths(ws, [18, 22, 7, 9, 14, 14, 14, 16, 16, 18, 16, 40, 26])

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH} with sheets: {wb.sheetnames}")
print(f"Overall Adjustments rows: {len(rows) - 1}")
