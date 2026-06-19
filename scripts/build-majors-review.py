"""
Build the manual-review workbook for the international-Major harvest:
  data-sources/majors-review.xlsx

Reads:
  data-sources/majors-new-teams.json   NEW + REVIEW teams (not in our Worlds DB)
  data-sources/majors-raw.json          every parsed Major team (for the audit tab)

Sheets:
  1. Leia-me           — instructions.
  2. New Teams         — one row PER PLAYER (+sub/coach) to review overalls. Seed
                         OVR = the player's peak overall already in our DB (blank =
                         brand-new player → estimate). Yellow = fill in.
  3. Skipped (audit)   — every team auto-excluded as a Worlds duplicate, so the
                         reviewer can confirm nothing real was dropped.

Regenerate after a fresh harvest:
  node scripts/fetch-liquipedia-majors.mjs
  python scripts/build-majors-review.py
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
NEW = os.path.join(ROOT, "data-sources", "majors-new-teams.json")
RAW = os.path.join(ROOT, "data-sources", "majors-raw.json")
OUT = os.path.join(ROOT, "data-sources", "majors-review.xlsx")

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F2A44")
INPUT_FILL = PatternFill("solid", fgColor="FFF3C4")
REVIEW_FILL = PatternFill("solid", fgColor="FDE2E2")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

new_teams = json.load(open(NEW, encoding="utf-8"))
raw = json.load(open(RAW, encoding="utf-8"))


def header(ws, cols, input_cols=()):
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = HEAD_FONT if i not in input_cols else Font(name=FONT, bold=True, color="7A5C00", size=11)
        c.fill = HEAD_FILL if i not in input_cols else INPUT_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


wb = Workbook()

# --- Leia-me ---
ws = wb.active
ws.title = "Leia-me"
ws.sheet_view.showGridLines = False
for i, (t, b, s) in enumerate([
    ("Majors internacionais — base nova para revisão", True, 15),
    ("", False, 10),
    ("Times que jogaram Majors internacionais (RLCS 2021-22 → 2025) mas NÃO foram ao Worlds.", False, 10),
    ("Os times de Worlds já estão na nossa base e foram excluídos automaticamente (aba 'Skipped' lista quais).", False, 10),
    ("", False, 10),
    ("Aba 'New Teams': uma linha por jogador. Preencha 'OVR final' (amarelo). Linhas rosa = status REVIEW", False, 10),
    ("(1 jogador do time já existe na nossa base naquela season — confirme se é time distinto ou duplicata).", False, 10),
    ("'Seed OVR' = pico do overall do jogador já na nossa base (qualquer season) — ajuste para ESTA season/carreira.", False, 10),
    ("Vazio em Seed OVR = jogador novo (estime por: resultado do time no Major + stats avançados).", False, 10),
    ("", False, 10),
    ("Aba 'Skipped (audit)': confira se nenhum time real foi excluído por engano (rebrand/alias).", False, 10),
], start=1):
    cell = ws.cell(row=i, column=1, value=t)
    cell.font = Font(name=FONT, bold=b, size=s, color="1F2A44" if b else "111827")
ws.column_dimensions["A"].width = 112

# --- New Teams (per-player rows) ---
ws = wb.create_sheet("New Teams")
cols = ["Status", "Major", "Year", "Team", "Região (preencher)", "Role", "Player",
        "Seed OVR (base)", "OVR final (REVISAR)", "Notas"]
header(ws, cols, input_cols=(5, 9, 10))
r = 2
for t in new_teams:
    rows = []
    for i, p in enumerate(t["players"], start=1):
        rows.append((f"P{i}", p["name"], p.get("seedOverall")))
    if t.get("sub"):
        rows.append(("Sub", t["sub"], t.get("subSeed")))
    if t.get("coach"):
        rows.append(("Coach", t["coach"], None))
    for role, name, seed in rows:
        vals = [t["status"].upper(), t["major"], t["year"], t["team"], "", role, name,
                seed if seed is not None else "", "", ""]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = BODY
            cell.border = BORDER
            if ci in (5, 9, 10):
                cell.fill = INPUT_FILL
            if t["status"] == "review":
                if ci in (1,):
                    cell.fill = REVIEW_FILL
        r += 1
for col, w in zip("ABCDEFGHIJ", [9, 26, 7, 24, 16, 7, 18, 12, 14, 30]):
    ws.column_dimensions[col].width = w

# --- Skipped (audit) ---
ws = wb.create_sheet("Skipped (audit)")
header(ws, ["Major", "Year", "Team (excluído = foi ao Worlds)", "Players (Liquipedia)"])
r = 2
# rebuild skipped from raw minus the new/review teams
new_keys = {(t["major"], t["team"]) for t in new_teams}
for title, M in raw.items():
    for t in M["teams"]:
        if (M["label"], t["team"]) in new_keys:
            continue
        vals = [M["label"], M["year"], t["team"], ", ".join(t["players"])]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = BODY
            cell.border = BORDER
        r += 1
for col, w in zip("ABCD", [26, 7, 32, 40]):
    ws.column_dimensions[col].width = w

wb.save(OUT)
players = sum(len(t["players"]) + (1 if t.get("sub") else 0) + (1 if t.get("coach") else 0) for t in new_teams)
print(f"Wrote {OUT}")
print(f"  New Teams: {len(new_teams)} teams, {players} review rows")
